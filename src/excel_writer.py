"""
Append parsed invoices to master Excel spreadsheet, then move the processed PDFs out of the inbox.

Behaviour:
- Creates the workbook on the first run with styled headers
- Appends a new row for each invoice on subsequent runs
- Auto-resizes columns to fit content
- Moves processed PDFs from "invoices/inbox" to "invoices/processed" with a timestamped filename
"""

from __future__ import annotations
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.extractor import Invoice

logger = logging.getLogger(__name__)

HEADERS = [
    "Invoice Number",
    "Invoice Date",
    "Vendor Name",
    "Total Amount",
    "Currency",
    "Line Items",
    "Source File",
    "Extraction Method",
    "Processed Timestamp",
]

CURRENCY_SYMBOLS = {
    "USD": '"$"#,##0.00',
    "GBP": '"£"#,##0.00',
    "EUR": '"€"#,##0.00',
    "JPY": '"¥"#,##0.00',
}

def _create_workbook(path: Path) -> Workbook:
    """Create a new workbook with styled headers."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Invoices"

    # Style headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.freeze_panes = "A2"  # Freeze header row
    ws.row_dimensions[1].height = 24  # Set header row height

    wb.save(path)
    logger.info(f"Created a new workbook at %s", path)
    return wb

def _open_or_create(path: Path) -> tuple[Workbook, Worksheet]:
    """Open existing workbook or create a new one if it doesn't exist."""
    if path.exists():
        wb = load_workbook(path)
        ws = wb['Invoices']
        logger.info(f"Opened existing workbook at %s", path)
    else:
        wb = _create_workbook(path)
        ws = wb['Invoices']
    return wb, ws

# -----Auto-resize columns based on content-----
def _autoresize_columns(ws: Worksheet) -> None:
    """ 
    Set each column width to roughly fit its longest cell
    """
    for col_idx, _ in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

# PUBLIC API
def append_invoices_to_excel(invoices: Iterable[Invoice], output_path: Path, ) -> int:
    """
    Append each invoice as a new row in the master workbook.
    Returns the number of rows apppended.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb, ws = _open_or_create(output_path)
    appended = 0
    for invoice in invoices:
        row_idx = ws.max_row + 1
        row = [
            invoice.invoice_number,
            invoice.invoice_date.isoformat() if invoice.invoice_date else "",
            invoice.vendor_name or "",
            float(invoice.total_amount),
            invoice.currency,
            "; ".join(f"{li.description}: {li.amount}" for li in invoice.line_items),
            invoice.source_file,
            invoice.extraction_method,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

        total_amount_cell = ws.cell(row=row_idx, column=4)
        total_amount_cell.number_format = CURRENCY_SYMBOLS.get(invoice.currency, '"$"#,##0.00"')
        appended += 1
        logger.info("Appended row %d: %s - %s %s",
                    row_idx, invoice.invoice_number,
                    invoice.total_amount, invoice.currency)
        
    _autoresize_columns(ws)
    wb.save(output_path)
    return appended
    
def move_processed_pdf(
        pdfs_paths: Iterable[Path],
        processed_dir: Path,
) -> list[Path]:
    """
    Move processed PDFs to the "processed" directory with a timestamped filename.
    Returns a list of new paths for the moved files.
    """
    processed_dir = Path(processed_dir)
    processed_dir.mkdir(parents=True, exist_ok=True)

    moved = []
    for pdf_path in pdfs_paths:
        dest = processed_dir / pdf_path.name
        if dest.exists():
            stem, suffix = dest.stem, dest.suffix
            ts = datetime.now().strftime("%Y%m%d%H%M%S")
            dest = processed_dir / f"{stem}_{ts}{suffix}"

        shutil.move(str(pdf_path), str(dest))
        moved.append(dest)
        logger.info("Moved processed PDF from %s to %s", pdf_path, dest)
    return moved

# CLI entry point for testing
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    from src.extractor import parse_invoice

    project_root = Path(__file__).parent.parent
    inbox = project_root / "invoices" / "inbox"
    processed = project_root / "invoices" / "processed"
    output = project_root / "output" / "master.xlsx"

    pdfs = sorted(inbox.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs in {inbox}")
        raise SystemExit(0)

    invoices = []           # ← list defined ONCE before the loop
    successful_pdfs = []
    failed = []

    for pdf in pdfs:
        try:
            invoices.append(parse_invoice(pdf))
            successful_pdfs.append(pdf)
        except Exception as e:
            failed.append((pdf.name, str(e)))

    if invoices:
        n = append_invoices_to_excel(invoices, output)
        print(f"Appended {n} invoice(s) to {output}")
        move_processed_pdf(successful_pdfs, processed)

    if failed:
        print("\nFailed extractions:")
        for name, err in failed:
            print(f"  {name}: {err}")